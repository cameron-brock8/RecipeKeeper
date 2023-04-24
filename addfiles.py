import sqlite3
from openpyxl import load_workbook

# Open a connection to the database
conn = sqlite3.connect('database.db')
cursor = conn.cursor()

cursor.execute("DROP TABLE IF EXISTS recipe")
cursor.execute("""
    CREATE TABLE recipe (
    id INT PRIMARY KEY, 
    name CHAR(20) UNIQUE,
    meal CHAR(20),
    difficulty INT,
    diet CHAR(20)
);
""")

cursor.execute("DROP TABLE IF EXISTS ingredients")
cursor.execute("""
   CREATE TABLE ingredients (
    id INTEGER PRIMARY KEY,
    list TEXT,
    FOREIGN KEY (id) REFERENCES recipe(id)
);
""")
               
cursor.execute("DROP TABLE IF EXISTS directions")
cursor.execute("""
   CREATE TABLE directions(
    id INT PRIMARY KEY,
    steps TEXT,
    time INT,
    FOREIGN KEY (id) REFERENCES recipe(id)
);
""")
               
cursor.execute("DROP TABLE IF EXISTS uses")
cursor.execute("""
   CREATE TABLE uses(
    name CHAR(20) UNIQUE,
    list TEXT,
    PRIMARY KEY(name, list),
    FOREIGN KEY (name) REFERENCES recipe(name),
    FOREIGN KEY (list) REFERENCES ingredients(list)
);
""")

# Read data from the Excel spreadsheet
workbook = load_workbook(filename='RecipeDatabase.xlsx')
worksheet = workbook.active
index = 0

# Iterate over each row of the DataFrame and insert it into the database
for id, name, meal, difficulty, diet, ingredients, steps, time in worksheet.iter_rows(min_row=2, values_only=True):

    cursor.execute("""
        INSERT INTO recipe (id, name, meal, difficulty, diet)
        VALUES (?, ?, ?, ?, ?)
    """, (id, name, meal, difficulty, diet))

    cursor.execute("""
        INSERT INTO ingredients (id, list)
        VALUES (?, ?)
    """, (id, ingredients))

    cursor.execute("""
        INSERT INTO directions (id, steps, time)
        VALUES (?, ?, ?)
    """, ( id, steps, time ))

    cursor.execute("""
        INSERT INTO uses (name, list)
        VALUES (?, ?)
    """, (name, ingredients) ) 

# Commit the changes and close the connection
conn.commit()
conn.close()
